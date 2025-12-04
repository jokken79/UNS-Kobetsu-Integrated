"""
Copy Excel sheets EXACTLY as they are - preserving ALL formatting.

This script copies each sheet from the original Excel file to a separate file
WITHOUT any modification, preserving:
- Colors (fill, font)
- Borders
- Merged cells
- Column widths
- Row heights
- Print settings
"""
import shutil
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from copy import copy

SOURCE_FILE = Path(r"D:\ExcelKobetsukeiyakusho.xlsx")
OUTPUT_DIR = Path(__file__).parent.parent / "templates" / "excel"

# Sheet index to output name mapping
SHEET_CONFIG = {
    0: "kobetsu_keiyakusho.xlsx",
    1: "tsuchisho.xlsx",
    2: "daicho.xlsx",
    3: "hakenmoto_daicho.xlsx",
    4: "shugyo_joken.xlsx",
    5: "keiyakusho.xlsx",
}


def copy_cell_completely(src_cell, tgt_cell):
    """Copy cell value and ALL styles."""
    # Copy value
    tgt_cell.value = src_cell.value

    # Copy styles
    if src_cell.has_style:
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.alignment = copy(src_cell.alignment)
        tgt_cell.number_format = src_cell.number_format
        tgt_cell.protection = copy(src_cell.protection)


def copy_sheet_exact(src_ws, tgt_ws):
    """Copy worksheet with ALL formatting preserved."""

    # Copy all cells
    for row in src_ws.iter_rows():
        for cell in row:
            tgt_cell = tgt_ws.cell(row=cell.row, column=cell.column)
            copy_cell_completely(cell, tgt_cell)

    # Copy column dimensions
    for col_letter, col_dim in src_ws.column_dimensions.items():
        tgt_ws.column_dimensions[col_letter].width = col_dim.width
        tgt_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        tgt_ws.column_dimensions[col_letter].outlineLevel = col_dim.outlineLevel
        tgt_ws.column_dimensions[col_letter].collapsed = col_dim.collapsed

    # Copy row dimensions
    for row_num, row_dim in src_ws.row_dimensions.items():
        tgt_ws.row_dimensions[row_num].height = row_dim.height
        tgt_ws.row_dimensions[row_num].hidden = row_dim.hidden
        tgt_ws.row_dimensions[row_num].outlineLevel = row_dim.outlineLevel
        tgt_ws.row_dimensions[row_num].collapsed = row_dim.collapsed

    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        tgt_ws.merge_cells(str(merged_range))

    # Copy print settings
    tgt_ws.print_area = src_ws.print_area
    tgt_ws.print_title_rows = src_ws.print_title_rows
    tgt_ws.print_title_cols = src_ws.print_title_cols

    # Copy page setup
    tgt_ws.page_setup.orientation = src_ws.page_setup.orientation
    tgt_ws.page_setup.paperSize = src_ws.page_setup.paperSize
    tgt_ws.page_setup.scale = src_ws.page_setup.scale
    tgt_ws.page_setup.fitToWidth = src_ws.page_setup.fitToWidth
    tgt_ws.page_setup.fitToHeight = src_ws.page_setup.fitToHeight

    # Copy page margins
    tgt_ws.page_margins.left = src_ws.page_margins.left
    tgt_ws.page_margins.right = src_ws.page_margins.right
    tgt_ws.page_margins.top = src_ws.page_margins.top
    tgt_ws.page_margins.bottom = src_ws.page_margins.bottom
    tgt_ws.page_margins.header = src_ws.page_margins.header
    tgt_ws.page_margins.footer = src_ws.page_margins.footer


def main():
    print("=" * 60)
    print("Copying Excel sheets with EXACT formatting")
    print("=" * 60)

    if not SOURCE_FILE.exists():
        print(f"ERROR: Source not found: {SOURCE_FILE}")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load source workbook
    print(f"\nLoading: {SOURCE_FILE}")
    src_wb = openpyxl.load_workbook(SOURCE_FILE, data_only=False)

    print(f"Found {len(src_wb.sheetnames)} sheets")

    for idx, sheet_name in enumerate(src_wb.sheetnames):
        if idx not in SHEET_CONFIG:
            print(f"\nSkipping sheet {idx}: {sheet_name}")
            continue

        output_name = SHEET_CONFIG[idx]
        print(f"\nCopying sheet {idx}: {sheet_name}")
        print(f"  Output: {output_name}")

        # Get source worksheet
        src_ws = src_wb.worksheets[idx]

        # Create new workbook
        tgt_wb = Workbook()
        tgt_ws = tgt_wb.active
        tgt_ws.title = sheet_name

        # Copy everything
        copy_sheet_exact(src_ws, tgt_ws)

        # Save
        output_path = OUTPUT_DIR / output_name
        tgt_wb.save(output_path)
        print(f"  Saved: {output_path}")
        print(f"  Size: {output_path.stat().st_size / 1024:.1f} KB")

        tgt_wb.close()

    src_wb.close()

    print("\n" + "=" * 60)
    print("Done! Templates now have EXACT formatting from original.")
    print("=" * 60)


if __name__ == "__main__":
    main()
