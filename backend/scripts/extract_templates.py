"""
Extract document templates from original Excel file - PRINT AREA ONLY.

This script:
1. Reads the original Excel file (D:\ExcelKobetsukeiyakusho.xlsx)
2. Extracts ONLY the print area of each document sheet
3. Preserves exact formatting (fonts, borders, colors, merged cells)
4. Creates templates that match exactly what would be printed

Usage:
    python scripts/extract_templates.py
"""
import re
import sys
from pathlib import Path
from copy import copy

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.page import PageMargins

# Configuration
SOURCE_FILE = r"D:\ExcelKobetsukeiyakusho.xlsx"
OUTPUT_DIR = Path(__file__).parent.parent / "templates" / "excel"

# Sheets to extract by index with their print areas
SHEET_INDEX_CONFIG = {
    0: {
        "output_name": "kobetsu_keiyakusho.xlsx",
        "description": "Individual Dispatch Contract",
        "print_area": "A1:AA64"
    },
    1: {
        "output_name": "tsuchisho.xlsx",
        "description": "Dispatch Notification",
        "print_area": "H1:P60"  # Starts at column H!
    },
    2: {
        "output_name": "daicho.xlsx",
        "description": "Individual Registry (DAICHO)",
        "print_area": "A1:BE78"
    },
    3: {
        "output_name": "hakenmoto_daicho.xlsx",
        "description": "Dispatch Origin Ledger",
        "print_area": "A1:AB71"
    },
    4: {
        "output_name": "shugyo_joken.xlsx",
        "description": "Employment Conditions",
        "print_area": "A1:AA56"
    },
    5: {
        "output_name": "keiyakusho.xlsx",
        "description": "Labor Contract",
        "print_area": "A1:R54"
    },
}


def parse_range(range_str: str) -> tuple:
    """Parse a range string like 'A1:AA64' into (min_col, min_row, max_col, max_row)."""
    # Remove sheet name if present
    if "!" in range_str:
        range_str = range_str.split("!")[-1]
    # Remove quotes and $ signs
    range_str = range_str.replace("'", "").replace("$", "")

    match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_str)
    if match:
        min_col = column_index_from_string(match.group(1))
        min_row = int(match.group(2))
        max_col = column_index_from_string(match.group(3))
        max_row = int(match.group(4))
        return min_col, min_row, max_col, max_row
    raise ValueError(f"Cannot parse range: {range_str}")


def copy_cell_style(source_cell, target_cell):
    """Copy all style properties from source to target cell."""
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


def extract_print_area(
    source_ws: Worksheet,
    data_ws: Worksheet,
    print_area: str,
) -> Workbook:
    """Extract only the print area from a worksheet."""
    min_col, min_row, max_col, max_row = parse_range(print_area)

    # Calculate offset (for sheets that don't start at A1)
    col_offset = min_col - 1
    row_offset = min_row - 1

    # Create new workbook
    template_wb = Workbook()
    template_ws = template_wb.active

    # Copy cells within print area only
    for src_row in range(min_row, max_row + 1):
        for src_col in range(min_col, max_col + 1):
            src_cell = source_ws.cell(row=src_row, column=src_col)

            # Target position (offset to start at A1)
            tgt_row = src_row - row_offset
            tgt_col = src_col - col_offset
            tgt_cell = template_ws.cell(row=tgt_row, column=tgt_col)

            # Copy value - prefer data_only version for formulas
            data_cell = data_ws.cell(row=src_row, column=src_col)
            if data_cell.value is not None:
                tgt_cell.value = data_cell.value
            elif src_cell.value is not None:
                if not (isinstance(src_cell.value, str) and src_cell.value.startswith("=")):
                    tgt_cell.value = src_cell.value

            # Copy style
            copy_cell_style(src_cell, tgt_cell)

    # Copy column widths (only for columns in print area)
    for src_col in range(min_col, max_col + 1):
        src_letter = get_column_letter(src_col)
        tgt_col = src_col - col_offset
        tgt_letter = get_column_letter(tgt_col)

        if src_letter in source_ws.column_dimensions:
            src_dim = source_ws.column_dimensions[src_letter]
            template_ws.column_dimensions[tgt_letter].width = src_dim.width
            if src_dim.hidden:
                template_ws.column_dimensions[tgt_letter].hidden = True

    # Copy row heights (only for rows in print area)
    for src_row in range(min_row, max_row + 1):
        tgt_row = src_row - row_offset
        if src_row in source_ws.row_dimensions:
            src_dim = source_ws.row_dimensions[src_row]
            if src_dim.height:
                template_ws.row_dimensions[tgt_row].height = src_dim.height
            if src_dim.hidden:
                template_ws.row_dimensions[tgt_row].hidden = True

    # Copy merged cells (only those within print area, with offset)
    for merged_range in source_ws.merged_cells.ranges:
        # Check if merge is within print area
        if (merged_range.min_col >= min_col and merged_range.max_col <= max_col and
            merged_range.min_row >= min_row and merged_range.max_row <= max_row):

            # Calculate new range with offset
            new_min_col = merged_range.min_col - col_offset
            new_max_col = merged_range.max_col - col_offset
            new_min_row = merged_range.min_row - row_offset
            new_max_row = merged_range.max_row - row_offset

            new_range = f"{get_column_letter(new_min_col)}{new_min_row}:{get_column_letter(new_max_col)}{new_max_row}"
            try:
                template_ws.merge_cells(new_range)
            except Exception:
                pass

    # Set print area for the new sheet (starts at A1)
    new_max_col = max_col - col_offset
    new_max_row = max_row - row_offset
    template_ws.print_area = f"A1:{get_column_letter(new_max_col)}{new_max_row}"

    # Copy page setup
    try:
        if source_ws.page_setup:
            template_ws.page_setup.orientation = source_ws.page_setup.orientation
            template_ws.page_setup.paperSize = source_ws.page_setup.paperSize
            template_ws.page_setup.fitToPage = source_ws.page_setup.fitToPage
            template_ws.page_setup.fitToWidth = source_ws.page_setup.fitToWidth
            template_ws.page_setup.fitToHeight = source_ws.page_setup.fitToHeight
    except Exception:
        pass

    # Copy page margins
    try:
        if source_ws.page_margins:
            template_ws.page_margins = PageMargins(
                left=source_ws.page_margins.left,
                right=source_ws.page_margins.right,
                top=source_ws.page_margins.top,
                bottom=source_ws.page_margins.bottom,
                header=source_ws.page_margins.header,
                footer=source_ws.page_margins.footer,
            )
    except Exception:
        pass

    return template_wb


def extract_all_templates():
    """Extract all document templates from the source Excel file."""
    if not Path(SOURCE_FILE).exists():
        print(f"ERROR: Source file not found: {SOURCE_FILE}")
        return False

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {OUTPUT_DIR}")

    # Load workbook with data_only=True to get calculated formula values
    print("Loading workbook with calculated values...")
    wb_data = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    # Also load with formulas to preserve formatting
    print("Loading workbook with formatting...")
    wb_format = openpyxl.load_workbook(SOURCE_FILE, data_only=False)

    print(f"\n{'=' * 60}")
    print("EXTRACTING PRINT AREAS ONLY")
    print("=" * 60)

    for idx, sheet_name in enumerate(wb_format.sheetnames):
        if idx not in SHEET_INDEX_CONFIG:
            continue

        config = SHEET_INDEX_CONFIG[idx]
        print_area = config["print_area"]

        print(f"\nSheet {idx}: {sheet_name}")
        print(f"  Print Area: {print_area}")
        print(f"  Output: {config['output_name']}")

        try:
            source_ws = wb_format[sheet_name]
            data_ws = wb_data[sheet_name]

            # Extract only print area
            template_wb = extract_print_area(source_ws, data_ws, print_area)

            # Save template
            output_path = OUTPUT_DIR / config["output_name"]
            template_wb.save(output_path)

            # Get dimensions of result
            result_ws = template_wb.active
            print(f"  Result: {result_ws.dimensions}")
            print(f"  Saved: {output_path}")

            template_wb.close()

        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback
            traceback.print_exc()

    wb_data.close()
    wb_format.close()

    print(f"\n{'=' * 60}")
    print("EXTRACTION COMPLETE")
    print("=" * 60)

    # List created files
    print("\nCreated templates:")
    for f in sorted(OUTPUT_DIR.glob("*.xlsx")):
        size = f.stat().st_size / 1024
        print(f"  {f.name} ({size:.1f} KB)")

    return True


def main():
    print("Excel Template Extraction Tool (Print Area Only)")
    print("=" * 60)
    print(f"Source: {SOURCE_FILE}")
    print(f"Output: {OUTPUT_DIR}")

    success = extract_all_templates()

    if success:
        print("\nTemplates extracted successfully!")
        print("Each template contains ONLY the printable area.")
    else:
        print("\nTemplate extraction failed.")
        sys.exit(1)


if __name__ == "__main__":
    main()
