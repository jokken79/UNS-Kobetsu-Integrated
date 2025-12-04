"""
Create clean templates with calculated values and all formatting.

This combines:
1. Formatting from the original (data_only=False)
2. Calculated values from formulas (data_only=True)

Result: Clean templates ready for data replacement.
"""
from pathlib import Path
from copy import copy
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SOURCE_FILE = Path(r"D:\ExcelKobetsukeiyakusho.xlsx")
OUTPUT_DIR = Path(__file__).parent.parent / "templates" / "excel"

SHEET_CONFIG = {
    0: "kobetsu_keiyakusho.xlsx",
    1: "tsuchisho.xlsx",
    2: "daicho.xlsx",
    3: "hakenmoto_daicho.xlsx",
    4: "shugyo_joken.xlsx",
    5: "keiyakusho.xlsx",
}


def copy_cell_style(src, tgt):
    """Copy all cell styles."""
    if src.has_style:
        tgt.font = copy(src.font)
        tgt.fill = copy(src.fill)
        tgt.border = copy(src.border)
        tgt.alignment = copy(src.alignment)
        tgt.number_format = src.number_format
        tgt.protection = copy(src.protection)


def create_template(format_ws, data_ws, output_path):
    """Create a clean template combining format and data."""
    wb = Workbook()
    ws = wb.active
    ws.title = format_ws.title

    # Get print area bounds
    print_area = format_ws.print_area
    if print_area:
        # Parse print area like "'Sheet'!$A$1:$AA$64"
        ws.print_area = print_area

    # Copy all cells (format from format_ws, value from data_ws)
    for row in format_ws.iter_rows():
        for cell in row:
            tgt = ws.cell(row=cell.row, column=cell.column)

            # Get value from data version (calculated formulas)
            data_cell = data_ws.cell(row=cell.row, column=cell.column)
            if data_cell.value is not None:
                tgt.value = data_cell.value
            elif cell.value is not None:
                # If no calculated value, check if it's a formula
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    tgt.value = None  # Skip formulas
                else:
                    tgt.value = cell.value

            # Copy style from format version
            copy_cell_style(cell, tgt)

    # Copy column widths
    for col_letter, dim in format_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width
        ws.column_dimensions[col_letter].hidden = dim.hidden

    # Copy row heights
    for row_num, dim in format_ws.row_dimensions.items():
        if dim.height:
            ws.row_dimensions[row_num].height = dim.height
        ws.row_dimensions[row_num].hidden = dim.hidden

    # Copy merged cells
    for merged in format_ws.merged_cells.ranges:
        ws.merge_cells(str(merged))

    # Copy page setup
    ws.page_setup.orientation = format_ws.page_setup.orientation
    ws.page_setup.paperSize = format_ws.page_setup.paperSize
    ws.page_setup.scale = format_ws.page_setup.scale
    ws.page_setup.fitToWidth = format_ws.page_setup.fitToWidth
    ws.page_setup.fitToHeight = format_ws.page_setup.fitToHeight

    # Copy margins
    ws.page_margins.left = format_ws.page_margins.left
    ws.page_margins.right = format_ws.page_margins.right
    ws.page_margins.top = format_ws.page_margins.top
    ws.page_margins.bottom = format_ws.page_margins.bottom

    # Save
    wb.save(output_path)
    wb.close()
    return output_path


def main():
    print("=" * 60)
    print("Creating Clean Templates (Format + Calculated Values)")
    print("=" * 60)

    if not SOURCE_FILE.exists():
        print(f"ERROR: Source not found: {SOURCE_FILE}")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load both versions
    print(f"\nLoading source with formulas (for formatting)...")
    wb_format = openpyxl.load_workbook(SOURCE_FILE, data_only=False)

    print(f"Loading source with data_only (for calculated values)...")
    wb_data = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    for idx, output_name in SHEET_CONFIG.items():
        print(f"\nSheet {idx}: {wb_format.sheetnames[idx]}")
        print(f"  Output: {output_name}")

        format_ws = wb_format.worksheets[idx]
        data_ws = wb_data.worksheets[idx]

        output_path = OUTPUT_DIR / output_name
        create_template(format_ws, data_ws, output_path)

        print(f"  Saved: {output_path}")
        print(f"  Size: {output_path.stat().st_size / 1024:.1f} KB")

    wb_format.close()
    wb_data.close()

    print("\n" + "=" * 60)
    print("Done! Templates have format + calculated values.")
    print("=" * 60)


if __name__ == "__main__":
    main()
