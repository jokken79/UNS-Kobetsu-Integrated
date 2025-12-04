"""
Verify Excel template formatting matches PDF requirements.
Compares template dimensions with expected A4 print area specifications.
"""
from pathlib import Path
import openpyxl

TEMPLATE_DIR = Path(__file__).parent.parent / "templates" / "excel"

# Expected specifications based on PDF analysis
PDF_SPECS = {
    "kobetsu_keiyakusho.xlsx": {
        "name": "人材派遣個別契約書",
        "orientation": "portrait",
        "columns": 27,  # A to AA
        "rows": 64,
        "paper_size": 9,  # A4
    },
    "tsuchisho.xlsx": {
        "name": "派遣先通知書",
        "orientation": "portrait",
        "columns": 9,  # H to P range
        "rows": 60,
        "paper_size": 9,
    },
    "daicho.xlsx": {
        "name": "派遣先管理台帳 (DAICHO)",
        "orientation": "portrait",
        "columns": 57,  # A to BE
        "rows": 78,
        "paper_size": 9,
    },
    "hakenmoto_daicho.xlsx": {
        "name": "派遣元管理台帳",
        "orientation": "portrait",
        "columns": 28,  # A to AB
        "rows": 71,
        "paper_size": 9,
    },
    "shugyo_joken.xlsx": {
        "name": "労働契約書兼就業条件明示書",
        "orientation": "landscape",  # LANDSCAPE!
        "columns": 27,  # A to AA
        "rows": 56,
        "paper_size": 9,
    },
    "keiyakusho.xlsx": {
        "name": "契約書",
        "orientation": "portrait",
        "columns": 18,  # A to R
        "rows": 54,
        "paper_size": 9,
    },
}


def verify_template(filename: str, spec: dict) -> dict:
    """Verify a single template against specifications."""
    filepath = TEMPLATE_DIR / filename

    if not filepath.exists():
        return {"error": f"Template not found: {filepath}"}

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    result = {
        "name": spec["name"],
        "file": filename,
        "status": "OK",
        "issues": [],
    }

    # Check dimensions
    dimensions = ws.dimensions
    if dimensions:
        result["actual_dimensions"] = dimensions
    else:
        result["actual_dimensions"] = "Unknown"

    # Check print area
    print_area = ws.print_area
    result["print_area"] = print_area

    # Check page setup
    page_setup = ws.page_setup
    actual_orientation = page_setup.orientation
    expected_orientation = spec["orientation"]

    result["orientation"] = {
        "expected": expected_orientation,
        "actual": actual_orientation,
    }

    if actual_orientation and actual_orientation.lower() != expected_orientation.lower():
        result["issues"].append(f"Orientation mismatch: expected {expected_orientation}, got {actual_orientation}")
        result["status"] = "WARNING"

    # Check paper size
    paper_size = page_setup.paperSize
    result["paper_size"] = {
        "expected": spec["paper_size"],
        "actual": paper_size,
    }

    if paper_size != spec["paper_size"]:
        result["issues"].append(f"Paper size mismatch: expected {spec['paper_size']}, got {paper_size}")
        result["status"] = "WARNING"

    # Check fit to page settings
    result["fit_settings"] = {
        "fitToWidth": page_setup.fitToWidth,
        "fitToHeight": page_setup.fitToHeight,
        "scale": page_setup.scale,
    }

    # Count rows and columns with content
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    result["content_range"] = {
        "expected_cols": spec["columns"],
        "actual_cols": max_col,
        "expected_rows": spec["rows"],
        "actual_rows": max_row,
    }

    # Count non-empty cells
    non_empty = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                non_empty += 1
    result["non_empty_cells"] = non_empty

    # Count merged cells
    merged_count = len(ws.merged_cells.ranges)
    result["merged_cells"] = merged_count

    wb.close()
    return result


def main():
    print("=" * 70)
    print("TEMPLATE FORMAT VERIFICATION")
    print("=" * 70)
    print(f"\nTemplate directory: {TEMPLATE_DIR}")
    print(f"Templates found: {list(TEMPLATE_DIR.glob('*.xlsx'))}")
    print()

    all_ok = True

    for filename, spec in PDF_SPECS.items():
        print(f"\n{'=' * 70}")
        print(f"Checking: {filename}")
        print(f"  Document: {spec['name']}")
        print(f"  Expected: {spec['columns']} cols x {spec['rows']} rows, {spec['orientation']}")
        print("-" * 70)

        result = verify_template(filename, spec)

        if "error" in result:
            print(f"  ERROR: {result['error']}")
            all_ok = False
            continue

        print(f"  Status: {result['status']}")
        print(f"  Actual dimensions: {result['actual_dimensions']}")
        print(f"  Print area: {result['print_area']}")
        print(f"  Orientation: {result['orientation']}")
        print(f"  Paper size: {result['paper_size']}")
        print(f"  Fit settings: {result['fit_settings']}")
        print(f"  Content range: {result['content_range']}")
        print(f"  Non-empty cells: {result['non_empty_cells']}")
        print(f"  Merged cells: {result['merged_cells']}")

        if result['issues']:
            print(f"  ISSUES:")
            for issue in result['issues']:
                print(f"    - {issue}")
            all_ok = False

    print(f"\n{'=' * 70}")
    if all_ok:
        print("ALL TEMPLATES VERIFIED OK")
    else:
        print("SOME TEMPLATES HAVE ISSUES - SEE ABOVE")
    print("=" * 70)


if __name__ == "__main__":
    main()
