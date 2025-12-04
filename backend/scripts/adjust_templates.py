"""
Adjust Excel templates to match exact A4 print format.

Based on PDF analysis:
- A4 Portrait: 210mm x 297mm (most documents)
- A4 Landscape: 297mm x 210mm (労働契約書 only)

Excel column width units: 1 unit ≈ 7 pixels ≈ 1.83mm at 96 DPI
Excel row height units: 1 point = 0.35mm
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

TEMPLATE_DIR = Path(__file__).parent.parent / "templates" / "excel"

# A4 dimensions in mm
A4_WIDTH_MM = 210
A4_HEIGHT_MM = 297

# Margin settings (in inches for openpyxl)
MM_TO_INCH = 0.0393701
MARGIN_TOP_MM = 10
MARGIN_BOTTOM_MM = 10
MARGIN_LEFT_MM = 10
MARGIN_RIGHT_MM = 10

# Document configurations based on PDF analysis
DOCUMENT_CONFIGS = {
    "kobetsu_keiyakusho.xlsx": {
        "name": "人材派遣個別契約書",
        "orientation": "portrait",
        "columns": 27,  # A to AA
        "rows": 64,
        "fit_to_width": 1,
        "fit_to_height": 1,
        "margins": {"top": 10, "bottom": 10, "left": 8, "right": 8},
    },
    "tsuchisho.xlsx": {
        "name": "派遣先通知書",
        "orientation": "portrait",
        "columns": 9,  # After offset from H1:P60
        "rows": 60,
        "fit_to_width": 1,
        "fit_to_height": 1,
        "margins": {"top": 15, "bottom": 10, "left": 15, "right": 15},
    },
    "daicho.xlsx": {
        "name": "派遣先管理台帳",
        "orientation": "portrait",
        "columns": 57,  # A to BE
        "rows": 78,
        "fit_to_width": 1,
        "fit_to_height": 1,
        "margins": {"top": 10, "bottom": 10, "left": 8, "right": 8},
    },
    "hakenmoto_daicho.xlsx": {
        "name": "派遣元管理台帳",
        "orientation": "portrait",
        "columns": 28,  # A to AB
        "rows": 71,
        "fit_to_width": 1,
        "fit_to_height": 1,
        "margins": {"top": 10, "bottom": 10, "left": 8, "right": 8},
    },
    "shugyo_joken.xlsx": {
        "name": "労働契約書兼就業条件明示書",
        "orientation": "landscape",  # LANDSCAPE!
        "columns": 27,  # A to AA
        "rows": 56,
        "fit_to_width": 1,
        "fit_to_height": 1,
        "margins": {"top": 8, "bottom": 8, "left": 10, "right": 10},
    },
    "keiyakusho.xlsx": {
        "name": "契約書",
        "orientation": "portrait",
        "columns": 18,  # A to R
        "rows": 54,
        "fit_to_width": 1,
        "fit_to_height": 1,
        "margins": {"top": 15, "bottom": 10, "left": 15, "right": 15},
    },
}


def adjust_template(filename: str, config: dict):
    """Adjust a single template to match A4 print specifications."""
    filepath = TEMPLATE_DIR / filename
    if not filepath.exists():
        print(f"  ERROR: File not found: {filepath}")
        return False

    print(f"\nAdjusting: {filename}")
    print(f"  Document: {config['name']}")
    print(f"  Orientation: {config['orientation']}")

    wb = load_workbook(filepath)
    ws = wb.active

    # Calculate printable area in mm
    margins = config["margins"]
    if config["orientation"] == "portrait":
        printable_width_mm = A4_WIDTH_MM - margins["left"] - margins["right"]
        printable_height_mm = A4_HEIGHT_MM - margins["top"] - margins["bottom"]
    else:  # landscape
        printable_width_mm = A4_HEIGHT_MM - margins["left"] - margins["right"]
        printable_height_mm = A4_WIDTH_MM - margins["top"] - margins["bottom"]

    print(f"  Printable area: {printable_width_mm}mm x {printable_height_mm}mm")

    # Set page orientation
    if config["orientation"] == "landscape":
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9  # A4
    else:
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = 9  # A4

    # Set fit to page (with error handling)
    try:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except (AttributeError, TypeError):
        pass

    try:
        ws.page_setup.fitToWidth = config.get("fit_to_width", 1)
        ws.page_setup.fitToHeight = config.get("fit_to_height", 1)
    except (AttributeError, TypeError):
        pass

    # Set margins (convert mm to inches)
    ws.page_margins = PageMargins(
        left=margins["left"] * MM_TO_INCH,
        right=margins["right"] * MM_TO_INCH,
        top=margins["top"] * MM_TO_INCH,
        bottom=margins["bottom"] * MM_TO_INCH,
        header=5 * MM_TO_INCH,
        footer=5 * MM_TO_INCH,
    )

    # Set print area
    max_col = config["columns"]
    max_row = config["rows"]
    print_area = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.print_area = print_area
    print(f"  Print area: {print_area}")

    # Calculate optimal column width
    # Excel column width: characters at default font
    # Approximate: 1 character width ≈ 7 pixels ≈ 1.85mm
    total_col_width_mm = printable_width_mm
    mm_per_column = total_col_width_mm / max_col
    # Convert mm to Excel units (approximately)
    # Excel width = mm / 2.3 (rough approximation)
    excel_width = mm_per_column / 2.3

    print(f"  Columns: {max_col}, Width per column: ~{excel_width:.1f} units")

    # Don't modify column widths if they're already set from template
    # Just ensure print settings are correct

    # Save
    wb.save(filepath)
    wb.close()
    print(f"  Saved: {filepath}")
    return True


def main():
    print("=" * 60)
    print("Adjusting Excel Templates for A4 Print")
    print("=" * 60)

    success_count = 0
    for filename, config in DOCUMENT_CONFIGS.items():
        if adjust_template(filename, config):
            success_count += 1

    print(f"\n{'=' * 60}")
    print(f"Adjusted {success_count}/{len(DOCUMENT_CONFIGS)} templates")
    print("=" * 60)


if __name__ == "__main__":
    main()
