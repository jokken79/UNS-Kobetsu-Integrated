"""
Template Manager - Excel template loading and caching.

This module provides efficient template management for Excel-based document generation.
Templates are loaded once and cached in memory for reuse.
"""
import copy
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from app.core.config import settings


class TemplateManager:
    """
    Manages Excel template loading with caching.

    Templates are loaded from EXCEL_TEMPLATE_DIR and cached in memory.
    Each get_template() call returns a deep copy to prevent modification
    of the cached template.
    """

    _cache: Dict[str, Workbook] = {}
    _initialized: bool = False

    # Template file mappings (romaji name -> filename)
    TEMPLATE_FILES = {
        "kobetsu_keiyakusho": "kobetsu_keiyakusho.xlsx",
        "tsuchisho": "tsuchisho.xlsx",
        "daicho": "daicho.xlsx",
        "hakenmoto_daicho": "hakenmoto_daicho.xlsx",
        "shugyo_joken": "shugyo_joken.xlsx",
        "keiyakusho": "keiyakusho.xlsx",
    }

    # Template descriptions (for documentation)
    TEMPLATE_DESCRIPTIONS = {
        "kobetsu_keiyakusho": "Individual Dispatch Contract (個別契約書)",
        "tsuchisho": "Dispatch Notification (通知書)",
        "daicho": "Individual Registry (DAICHO/台帳)",
        "hakenmoto_daicho": "Dispatch Origin Ledger (派遣元管理台帳)",
        "shugyo_joken": "Employment Conditions Document (就業条件明示書)",
        "keiyakusho": "Labor Contract (契約書)",
    }

    @classmethod
    def get_template_dir(cls) -> Path:
        """Get the template directory path."""
        template_dir = getattr(settings, 'EXCEL_TEMPLATE_DIR', './templates/excel')
        return Path(template_dir)

    @classmethod
    def get_template(cls, template_name: str) -> Workbook:
        """
        Get a fresh copy of a template workbook.

        Args:
            template_name: Name of the template (e.g., 'kobetsu_keiyakusho')

        Returns:
            A fresh load of the template Workbook

        Raises:
            ValueError: If template name is not recognized
            FileNotFoundError: If template file doesn't exist
        """
        # Load fresh copy each time (deepcopy doesn't work with openpyxl styles)
        filename = cls.TEMPLATE_FILES.get(template_name)
        if not filename:
            available = ", ".join(cls.TEMPLATE_FILES.keys())
            raise ValueError(
                f"Unknown template: '{template_name}'. "
                f"Available templates: {available}"
            )

        template_dir = cls.get_template_dir()
        template_path = template_dir / filename

        if not template_path.exists():
            raise FileNotFoundError(
                f"Template file not found: {template_path}. "
                f"Please run extract_templates.py first."
            )

        return load_workbook(template_path)

    @classmethod
    def _load_template(cls, template_name: str) -> None:
        """Load template into cache."""
        filename = cls.TEMPLATE_FILES.get(template_name)
        if not filename:
            available = ", ".join(cls.TEMPLATE_FILES.keys())
            raise ValueError(
                f"Unknown template: '{template_name}'. "
                f"Available templates: {available}"
            )

        template_dir = cls.get_template_dir()
        template_path = template_dir / filename

        if not template_path.exists():
            raise FileNotFoundError(
                f"Template file not found: {template_path}. "
                f"Please run extract_templates.py first."
            )

        cls._cache[template_name] = load_workbook(template_path)

    @classmethod
    def clear_cache(cls) -> None:
        """Clear the template cache (for testing or hot-reload)."""
        for wb in cls._cache.values():
            try:
                wb.close()
            except Exception:
                pass
        cls._cache.clear()

    @classmethod
    def preload_all(cls) -> Dict[str, bool]:
        """
        Preload all templates into cache.

        Returns:
            Dict mapping template name to success status
        """
        results = {}
        for name in cls.TEMPLATE_FILES.keys():
            try:
                cls._load_template(name)
                results[name] = True
            except Exception as e:
                results[name] = False
                print(f"Failed to load template '{name}': {e}")
        return results

    @classmethod
    def list_templates(cls) -> Dict[str, dict]:
        """
        List all available templates with metadata.

        Returns:
            Dict mapping template name to info dict with:
            - filename: The template file name
            - description: Human-readable description
            - exists: Whether the template file exists
            - cached: Whether template is currently cached
        """
        template_dir = cls.get_template_dir()
        result = {}

        for name, filename in cls.TEMPLATE_FILES.items():
            path = template_dir / filename
            result[name] = {
                "filename": filename,
                "description": cls.TEMPLATE_DESCRIPTIONS.get(name, ""),
                "exists": path.exists(),
                "cached": name in cls._cache,
                "size_kb": round(path.stat().st_size / 1024, 1) if path.exists() else 0,
            }

        return result

    @classmethod
    def validate_templates(cls) -> tuple[bool, list[str]]:
        """
        Validate that all required templates exist.

        Returns:
            Tuple of (all_valid, list_of_missing_templates)
        """
        template_dir = cls.get_template_dir()
        missing = []

        for name, filename in cls.TEMPLATE_FILES.items():
            path = template_dir / filename
            if not path.exists():
                missing.append(name)

        return (len(missing) == 0, missing)


# Convenience function
def get_template(template_name: str) -> Workbook:
    """Convenience function to get a template."""
    return TemplateManager.get_template(template_name)
