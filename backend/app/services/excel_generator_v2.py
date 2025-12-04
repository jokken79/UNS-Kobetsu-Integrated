"""
Excel Generator V2 - Template-based generation using JSON data.

This is the NEW approach: JSON → Excel Template → Filled Document

Advantages:
- Perfect formatting preservation
- Easy to maintain (edit templates in Excel)
- Fast generation
- Compatible with original Excel system
"""
from datetime import date, time
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Optional, Union

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import settings
from app.schemas.document_data import DocumentDataSchema


class ExcelGeneratorV2:
    """
    Excel document generator using templates and JSON data.

    This generator:
    1. Loads a pre-formatted Excel template
    2. Fills it with data from JSON (DocumentDataSchema)
    3. Returns the completed document as bytes

    Templates are stored in: backend/templates/excel/
    """

    def __init__(self, data: DocumentDataSchema):
        """
        Initialize generator with document data.

        Args:
            data: DocumentDataSchema (JSON format)
        """
        self.data = data
        self.template_dir = Path(settings.EXCEL_TEMPLATE_DIR)

    # ========================================
    # FORMATTING HELPERS
    # ========================================

    def _format_date_japanese(self, d: Optional[date]) -> str:
        """Format date as Japanese era (令和X年X月X日)."""
        if d is None:
            return ""

        # Reiwa era started May 1, 2019
        if d.year >= 2019 and (d.month > 4 or (d.month == 4 and d.year > 2019)):
            reiwa_year = d.year - 2018
            return f"令和{reiwa_year}年{d.month}月{d.day}日"

        # Heisei era
        if d.year >= 1989:
            heisei_year = d.year - 1988
            return f"平成{heisei_year}年{d.month}月{d.day}日"

        return f"{d.year}年{d.month}月{d.day}日"

    def _format_time(self, t: Optional[time]) -> str:
        """Format time as HH:MM."""
        if t is None:
            return ""
        return t.strftime("%H:%M")

    def _format_time_range(self, start: Optional[time], end: Optional[time]) -> str:
        """Format time range as HH:MM～HH:MM."""
        start_str = self._format_time(start)
        end_str = self._format_time(end)
        if start_str and end_str:
            return f"{start_str}～{end_str}"
        return ""

    def _format_currency(self, value: Optional[Union[Decimal, int, float]]) -> str:
        """Format as Japanese yen (¥X,XXX)."""
        if value is None:
            return ""
        return f"¥{int(value):,}"

    def _format_work_days(self, days: list) -> str:
        """Format work days list as string."""
        if not days:
            return "月・火・水・木・金"
        return "・".join(days)

    # ========================================
    # TEMPLATE OPERATIONS
    # ========================================

    def _load_template(self, template_name: str) -> Workbook:
        """Load Excel template from disk."""
        template_path = self.template_dir / f"{template_name}.xlsx"

        if not template_path.exists():
            raise FileNotFoundError(
                f"Template not found: {template_path}\n"
                f"Please run: python backend/scripts/extract_templates.py"
            )

        return load_workbook(template_path)

    def _replace_placeholders(self, ws: Worksheet, replacements: dict) -> None:
        """
        Replace placeholders in worksheet with actual values.

        Searches for {{placeholder}} patterns and replaces them.
        """
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for placeholder, value in replacements.items():
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(
                                placeholder,
                                str(value) if value is not None else ""
                            )

    def _workbook_to_bytes(self, wb: Workbook) -> bytes:
        """Convert workbook to bytes for download."""
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ========================================
    # DOCUMENT GENERATORS
    # ========================================

    def generate_kobetsu_keiyakusho(self) -> bytes:
        """
        Generate 個別契約書 (Individual Dispatch Contract).

        Returns:
            Excel file as bytes
        """
        wb = self._load_template("kobetsu_keiyakusho")
        ws = wb.active

        # Build replacement dictionary from JSON data
        replacements = {
            # Contract info
            "{{contract_number}}": self.data.contract_number,
            "{{contract_date}}": self._format_date_japanese(self.data.contract_date),
            "{{document_date}}": self._format_date_japanese(self.data.document_date),

            # Dispatch period
            "{{dispatch_start}}": self._format_date_japanese(self.data.dispatch_start_date),
            "{{dispatch_end}}": self._format_date_japanese(self.data.dispatch_end_date),

            # Dispatch company (派遣元 - UNS Kikaku)
            "{{dispatch_company}}": self.data.dispatch_company.name,
            "{{dispatch_company_address}}": self.data.dispatch_company.address,
            "{{dispatch_company_tel}}": self.data.dispatch_company.tel,
            "{{license_number}}": self.data.dispatch_company.license_number or "",

            # Client company (派遣先)
            "{{client_company}}": self.data.client_company.name,
            "{{client_address}}": self.data.client_company.address,
            "{{client_tel}}": self.data.client_company.tel,

            # Worksite (就業場所)
            "{{worksite_name}}": self.data.worksite_name,
            "{{worksite_address}}": self.data.worksite_address,
            "{{organizational_unit}}": self.data.organizational_unit or "",

            # Work content (業務内容)
            "{{work_content}}": self.data.work_content,
            "{{responsibility}}": self.data.responsibility_level,

            # Supervisor (指揮命令者)
            "{{supervisor_name}}": self.data.supervisor.name,
            "{{supervisor_dept}}": self.data.supervisor.department or "",
            "{{supervisor_position}}": self.data.supervisor.position or "",
            "{{supervisor}}": self._format_supervisor(self.data.supervisor),

            # Work schedule (就業時間)
            "{{work_days}}": self._format_work_days(self.data.schedule.work_days),
            "{{work_start}}": self._format_time(self.data.schedule.work_start_time),
            "{{work_end}}": self._format_time(self.data.schedule.work_end_time),
            "{{work_time}}": self._format_time_range(
                self.data.schedule.work_start_time,
                self.data.schedule.work_end_time
            ),
            "{{break_minutes}}": f"{self.data.schedule.break_time_minutes}分",

            # Overtime (時間外労働)
            "{{overtime_day}}": f"{self.data.schedule.overtime_max_hours_day}時間/日",
            "{{overtime_month}}": f"{self.data.schedule.overtime_max_hours_month}時間/月",

            # Rates (派遣料金)
            "{{hourly_rate}}": self._format_currency(self.data.rates.hourly_rate),
            "{{overtime_rate}}": self._format_currency(self.data.rates.overtime_rate),
            "{{holiday_rate}}": self._format_currency(self.data.rates.holiday_rate),
            "{{night_rate}}": self._format_currency(self.data.rates.night_shift_rate),

            # Number of workers
            "{{num_workers}}": f"{self.data.number_of_workers}名",

            # Managers (責任者)
            "{{dispatch_manager}}": self.data.dispatch_manager.name,
            "{{dispatch_manager_phone}}": self.data.dispatch_manager.phone or "",
            "{{client_manager}}": self.data.client_manager.name,
            "{{client_manager_phone}}": self.data.client_manager.phone or "",

            # Complaint handlers (苦情処理)
            "{{dispatch_complaint_name}}": self.data.dispatch_complaint_handler.name,
            "{{dispatch_complaint_phone}}": self.data.dispatch_complaint_handler.phone or "",
            "{{client_complaint_name}}": self.data.client_complaint_handler.name,
            "{{client_complaint_phone}}": self.data.client_complaint_handler.phone or "",

            # Safety & welfare
            "{{safety_measures}}": self.data.safety_measures,
            "{{welfare}}": "・".join(self.data.welfare_facilities),

            # Legal checkboxes
            "{{kyotei_check}}": "☑" if self.data.is_kyotei_taisho else "☐",
            "{{mukeiko_check}}": "☑" if self.data.is_mukeiko_60over_only else "☐",
            "{{direct_hire_check}}": "☑" if self.data.is_direct_hire_prevention else "☐",
        }

        # Add employee data
        for i, emp in enumerate(self.data.employees, start=1):
            replacements[f"{{{{employee_{i}_name}}}}"] = emp.full_name
            replacements[f"{{{{employee_{i}_kana}}}}"] = emp.full_name_kana
            replacements[f"{{{{employee_{i}_number}}}}"] = emp.employee_number

        # Apply replacements
        self._replace_placeholders(ws, replacements)

        return self._workbook_to_bytes(wb)

    def _format_supervisor(self, person) -> str:
        """Format supervisor as single string."""
        parts = []
        if person.department:
            parts.append(person.department)
        if person.position:
            parts.append(person.position)
        if person.name:
            parts.append(person.name)
        return " ".join(parts)

    def generate_tsuchisho(self) -> bytes:
        """Generate 通知書 (Notification)."""
        wb = self._load_template("tsuchisho")
        ws = wb.active

        replacements = {
            "{{notification_date}}": self._format_date_japanese(self.data.document_date),
            "{{client_company}}": self.data.client_company.name,
            "{{dispatch_company}}": self.data.dispatch_company.name,
            "{{dispatch_address}}": self.data.dispatch_company.address,
            "{{license_number}}": self.data.dispatch_company.license_number or "",
            "{{dispatch_start}}": self._format_date_japanese(self.data.dispatch_start_date),
            "{{dispatch_end}}": self._format_date_japanese(self.data.dispatch_end_date),
            "{{worksite_name}}": self.data.worksite_name,
            "{{work_content}}": self.data.work_content,
            "{{work_time}}": self._format_time_range(
                self.data.schedule.work_start_time,
                self.data.schedule.work_end_time
            ),
        }

        # Add employees
        for i, emp in enumerate(self.data.employees, start=1):
            replacements[f"{{{{emp_{i}_name}}}}"] = emp.full_name
            replacements[f"{{{{emp_{i}_kana}}}}"] = emp.full_name_kana

        self._replace_placeholders(ws, replacements)
        return self._workbook_to_bytes(wb)

    def generate_all(self) -> dict:
        """
        Generate all available documents.

        Returns:
            Dict mapping document name to bytes
        """
        results = {}

        try:
            results["kobetsu_keiyakusho"] = self.generate_kobetsu_keiyakusho()
        except Exception as e:
            results["kobetsu_keiyakusho_error"] = str(e)

        try:
            results["tsuchisho"] = self.generate_tsuchisho()
        except Exception as e:
            results["tsuchisho_error"] = str(e)

        return results
