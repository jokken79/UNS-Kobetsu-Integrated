"""
Import Service - データインポートサービス

Handles importing factories and employees from JSON and Excel files.
Provides preview/validation before actual import.
Automatically links employees to factories and lines during import.
"""
import json
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.factory import Factory, FactoryLine
from app.models.employee import Employee


# Manual mapping for employee company_name -> (factory_company, factory_plant)
# This handles cases where automatic matching would fail
EMPLOYEE_TO_FACTORY_MAPPING = {
    # 高雄工業 variations
    "高雄工業 岡山": ("高雄工業株式会社", "岡山工場"),
    "高雄工業 本社": ("高雄工業株式会社", "本社工場"),
    "高雄工業 海南第一": ("高雄工業株式会社", "海南第一工場"),
    "高雄工業 海南第二": ("高雄工業株式会社", "海南第二工場"),
    "高雄工業 静岡": ("高雄工業株式会社", "静岡工場"),

    # 加藤木材工業
    "加藤木材工業 本社": ("加藤木材工業株式会社", "本社工場"),
    "加藤木材工業 春日井": ("加藤木材工業株式会社", "春日井工場"),

    # ユアサ工機
    "ユアサ工機 新城": ("ユアサ工機株式会社", "新城工場"),
    "ユアサ工機 御津": ("ユアサ工機株式会社", "本社工場"),

    # Simple matches
    "瑞陵精機": ("瑞陵精機株式会社", "恵那工場"),
    "三幸技研": ("三幸技研株式会社", "本社工場"),
    "六甲電子": ("六甲電子株式会社", "本社工場"),
    "川原鉄工所": ("株式会社川原鉄工所", "本社工場"),
    "オーツカ": ("株式会社オーツカ", "関ケ原工場"),
    "ピーエムアイ": ("ピーエムアイ有限会社", "本社工場"),
    "セイビテック": ("セイビテック株式会社", ""),

    # Half-width katakana variations
    "ﾃｨｰｹｰｴﾝｼﾞﾆｱﾘﾝｸﾞ": ("ティーケーエンジニアリング株式会社", "海南第二工場"),
    "ﾌｪﾆﾃｯｸｾﾐｺﾝﾀﾞｸﾀｰ 岡山": ("フェニテックセミコンダクター(株)", "鹿児島工場"),

    # 美鈴工業
    "美鈴工業 本社": ("株式会社美鈴工業", "本社工場"),
    "美鈴工業 本庄": ("株式会社美鈴工業", "本社工場"),

    # PATEC
    "PATEC": ("PATEC株式会社", "防府工場"),

    # コーリツ
    "コーリツ 本社": ("株式会社コーリツ", "本社工場"),
    "コーリツ 乙川": ("株式会社コーリツ", "乙川工場"),
    "コーリツ 亀崎": ("株式会社コーリツ", "亀崎工場"),
    "コーリツ 州の崎": ("株式会社コーリツ", "州の崎工場"),

    # プレテック
    "プレテック": ("プレテック株式会社", "本社工場"),

    # ワーク
    "ワーク 堺": ("株式会社ワーク", "堺工場"),
    "ワーク 志紀": ("株式会社ワーク", "志紀工場"),
}


class ImportValidationError:
    """Represents a validation error during import."""

    def __init__(self, row: int, field: str, message: str, value: Any = None):
        self.row = row
        self.field = field
        self.message = message
        self.value = value

    def to_dict(self) -> dict:
        return {
            "row": self.row,
            "field": self.field,
            "message": self.message,
            "value": str(self.value) if self.value is not None else None
        }


class ImportResult:
    """Result of an import operation."""

    def __init__(self):
        self.success = False
        self.total_rows = 0
        self.imported_count = 0
        self.updated_count = 0
        self.skipped_count = 0
        self.errors: List[ImportValidationError] = []
        self.preview_data: List[dict] = []
        self.message = ""

    def to_dict(self) -> dict:
        return {
            "success": self.success,
            "total_rows": self.total_rows,
            "imported_count": self.imported_count,
            "updated_count": self.updated_count,
            "skipped_count": self.skipped_count,
            "errors": [e.to_dict() for e in self.errors],
            "preview_data": self.preview_data[:100],  # Limit preview
            "message": self.message
        }


class ImportService:
    """Service for importing data from files."""

    def __init__(self, db: Session):
        self.db = db

    # ========================================
    # FACTORY/LINE LOOKUP (for employee linking)
    # ========================================

    def _normalize_company_name(self, name: str) -> str:
        """Normalize company name for matching."""
        if not name:
            return ""
        # Remove common suffixes
        name = re.sub(r'株式会社|有限会社|\(株\)|（株）', '', name)
        # Remove whitespace
        name = name.strip()
        return name

    def _find_factory_for_employee(self, company_name: str) -> Optional[Factory]:
        """
        Find matching factory for employee company_name.

        Uses manual mapping first, then fuzzy matching.
        """
        if not company_name:
            return None

        company_name = company_name.strip()

        # Check manual mapping first
        if company_name in EMPLOYEE_TO_FACTORY_MAPPING:
            company, plant = EMPLOYEE_TO_FACTORY_MAPPING[company_name]
            normalized_company = self._normalize_company_name(company)

            # Try exact plant match
            factory = self.db.query(Factory).filter(
                Factory.company_name.contains(normalized_company),
                Factory.plant_name == plant
            ).first()

            if not factory and plant:
                # Try without plant filter
                factory = self.db.query(Factory).filter(
                    Factory.company_name.contains(normalized_company)
                ).first()

            return factory

        # Fuzzy matching
        normalized = self._normalize_company_name(company_name)

        # Split by space to get company and plant parts
        parts = company_name.split()
        if len(parts) >= 2:
            company_part = parts[0]
            plant_part = parts[1] if len(parts) > 1 else ""

            # Try to find factory matching both parts
            factory = self.db.query(Factory).filter(
                Factory.company_name.contains(company_part),
                Factory.plant_name.contains(plant_part)
            ).first()

            if factory:
                return factory

        # Last resort: just match company name
        factory = self.db.query(Factory).filter(
            Factory.company_name.contains(normalized)
        ).first()

        return factory

    def _find_factory_line_for_employee(
        self, factory_id: int, department: str, line_name: str
    ) -> Optional[FactoryLine]:
        """
        Find matching factory line based on department and line_name.

        Priority:
        1. Exact match (department + line_name)
        2. Line name match only
        3. Department match only
        4. First available line
        """
        if not factory_id:
            return None

        query = self.db.query(FactoryLine).filter(FactoryLine.factory_id == factory_id)

        # Try exact match first
        if department and line_name:
            line = query.filter(
                FactoryLine.department == department,
                FactoryLine.line_name == line_name
            ).first()
            if line:
                return line

        # Try line_name match
        if line_name:
            line = query.filter(FactoryLine.line_name == line_name).first()
            if line:
                return line

        # Try department match
        if department:
            line = query.filter(FactoryLine.department == department).first()
            if line:
                return line

        # Return first available line if no match
        return query.first()

    def _link_employee_to_factory(
        self, employee: Employee, company_name: str, department: str, line_name: str
    ) -> Tuple[Optional[int], Optional[int]]:
        """
        Link employee to factory and factory_line based on their assignment info.

        Returns:
            (factory_id, factory_line_id) or (None, None) if no match
        """
        factory = self._find_factory_for_employee(company_name)
        if not factory:
            return None, None

        factory_line = self._find_factory_line_for_employee(
            factory.id, department, line_name
        )

        return factory.id, factory_line.id if factory_line else None

    # ========================================
    # FACTORY IMPORT
    # ========================================

    def preview_factories_json(self, content: bytes) -> ImportResult:
        """Preview factory data from JSON file."""
        result = ImportResult()

        try:
            data = json.loads(content.decode('utf-8'))

            # Handle both single object and array
            if isinstance(data, dict):
                factories = [data]
            else:
                factories = data

            result.total_rows = len(factories)

            for idx, factory_data in enumerate(factories, 1):
                validated, errors = self._validate_factory(idx, factory_data)
                result.errors.extend(errors)

                # Add to preview with validation status
                preview_item = {
                    "row": idx,
                    "company_name": factory_data.get("company_name", factory_data.get("派遣先名", "")),
                    "plant_name": factory_data.get("plant_name", factory_data.get("工場名", "")),
                    "conflict_date": factory_data.get("conflict_date", factory_data.get("抵触日", "")),
                    "is_valid": len(errors) == 0,
                    "errors": [e.message for e in errors],
                    "_raw": factory_data
                }
                result.preview_data.append(preview_item)

            result.success = len(result.errors) == 0
            result.message = f"{result.total_rows}件の工場データを読み込みました" if result.success else f"{len(result.errors)}件のエラーがあります"

        except json.JSONDecodeError as e:
            result.errors.append(ImportValidationError(0, "file", f"JSON解析エラー: {str(e)}"))
            result.message = "JSONファイルの形式が正しくありません"
        except Exception as e:
            result.errors.append(ImportValidationError(0, "file", f"エラー: {str(e)}"))
            result.message = f"ファイル読み込みエラー: {str(e)}"

        return result

    def preview_factories_excel(self, content: bytes) -> ImportResult:
        """Preview factory data from Excel file."""
        result = ImportResult()

        try:
            df = pd.read_excel(BytesIO(content), engine='openpyxl')
            result.total_rows = len(df)

            # Map Japanese column names to English
            column_mapping = {
                '派遣先名': 'company_name',
                '会社名': 'company_name',
                '工場名': 'plant_name',
                '工場住所': 'plant_address',
                '抵触日': 'conflict_date',
                '派遣先責任者': 'client_responsible_name',
                '派遣先苦情担当者': 'client_complaint_name',
                '締め日': 'closing_date',
                '支払日': 'payment_date',
            }

            df = df.rename(columns=column_mapping)

            for idx, row in df.iterrows():
                row_num = idx + 2  # Excel row (1-indexed + header)
                factory_data = row.to_dict()

                # Clean NaN values
                factory_data = {k: (None if pd.isna(v) else v) for k, v in factory_data.items()}

                validated, errors = self._validate_factory(row_num, factory_data)
                result.errors.extend(errors)

                preview_item = {
                    "row": row_num,
                    "company_name": factory_data.get("company_name", ""),
                    "plant_name": factory_data.get("plant_name", ""),
                    "conflict_date": str(factory_data.get("conflict_date", "")) if factory_data.get("conflict_date") else "",
                    "is_valid": len([e for e in errors if e.row == row_num]) == 0,
                    "errors": [e.message for e in errors if e.row == row_num],
                    "_raw": factory_data
                }
                result.preview_data.append(preview_item)

            result.success = len(result.errors) == 0
            result.message = f"{result.total_rows}件の工場データを読み込みました" if result.success else f"{len(result.errors)}件のエラーがあります"

        except Exception as e:
            result.errors.append(ImportValidationError(0, "file", f"Excelエラー: {str(e)}"))
            result.message = f"Excelファイル読み込みエラー: {str(e)}"

        return result

    def import_factories(self, preview_data: List[dict], mode: str = "create") -> ImportResult:
        """
        Actually import factories after preview confirmation.

        Args:
            preview_data: Validated data from preview
            mode: "create" (skip existing), "update" (update existing), "sync" (delete missing)
        """
        result = ImportResult()
        result.total_rows = len(preview_data)

        try:
            for item in preview_data:
                if not item.get("is_valid", False):
                    result.skipped_count += 1
                    continue

                raw_data = item.get("_raw", {})
                factory_id = f"{raw_data.get('company_name', '')}_{raw_data.get('plant_name', '')}".replace(" ", "_")

                # Check if exists
                existing = self.db.query(Factory).filter(Factory.factory_id == factory_id).first()

                if existing:
                    if mode in ["update", "sync"]:
                        # Update existing
                        self._update_factory(existing, raw_data)
                        result.updated_count += 1
                    else:
                        result.skipped_count += 1
                else:
                    # Create new
                    factory = self._create_factory(factory_id, raw_data)
                    self.db.add(factory)
                    result.imported_count += 1

            self.db.commit()
            result.success = True
            result.message = f"インポート完了: 新規{result.imported_count}件, 更新{result.updated_count}件, スキップ{result.skipped_count}件"

        except Exception as e:
            self.db.rollback()
            result.errors.append(ImportValidationError(0, "import", f"インポートエラー: {str(e)}"))
            result.message = f"インポート失敗: {str(e)}"

        return result

    def _validate_factory(self, row: int, data: dict) -> Tuple[dict, List[ImportValidationError]]:
        """Validate a single factory record."""
        errors = []

        company_name = data.get("company_name") or data.get("派遣先名")
        plant_name = data.get("plant_name") or data.get("工場名")

        if not company_name:
            errors.append(ImportValidationError(row, "company_name", "派遣先名は必須です"))

        if not plant_name:
            errors.append(ImportValidationError(row, "plant_name", "工場名は必須です"))

        # Validate conflict_date if present
        conflict_date = data.get("conflict_date") or data.get("抵触日")
        if conflict_date:
            try:
                if isinstance(conflict_date, str):
                    datetime.strptime(conflict_date, "%Y-%m-%d")
            except ValueError:
                errors.append(ImportValidationError(row, "conflict_date", "抵触日の形式が正しくありません (YYYY-MM-DD)", conflict_date))

        return data, errors

    def _create_factory(self, factory_id: str, data: dict) -> Factory:
        """Create a new Factory from import data."""
        conflict_date = data.get("conflict_date") or data.get("抵触日")
        if isinstance(conflict_date, str):
            conflict_date = datetime.strptime(conflict_date, "%Y-%m-%d").date()
        elif isinstance(conflict_date, datetime):
            conflict_date = conflict_date.date()

        return Factory(
            factory_id=factory_id,
            company_name=data.get("company_name") or data.get("派遣先名", ""),
            company_address=data.get("company_address") or data.get("派遣先住所"),
            company_phone=data.get("company_phone") or data.get("派遣先電話"),
            plant_name=data.get("plant_name") or data.get("工場名", ""),
            plant_address=data.get("plant_address") or data.get("工場住所"),
            plant_phone=data.get("plant_phone") or data.get("工場電話"),
            client_responsible_name=data.get("client_responsible_name") or data.get("派遣先責任者"),
            client_responsible_department=data.get("client_responsible_department") or data.get("派遣先責任者部署"),
            client_complaint_name=data.get("client_complaint_name") or data.get("派遣先苦情担当者"),
            client_complaint_department=data.get("client_complaint_department") or data.get("派遣先苦情担当部署"),
            dispatch_responsible_name=data.get("dispatch_responsible_name") or data.get("派遣元責任者"),
            dispatch_complaint_name=data.get("dispatch_complaint_name") or data.get("派遣元苦情担当者"),
            conflict_date=conflict_date,
            closing_date=data.get("closing_date") or data.get("締め日"),
            payment_date=data.get("payment_date") or data.get("支払日"),
            break_minutes=int(data.get("break_minutes", 60)),
            is_active=True
        )

    def _update_factory(self, factory: Factory, data: dict):
        """Update existing factory with new data."""
        if data.get("company_address") or data.get("派遣先住所"):
            factory.company_address = data.get("company_address") or data.get("派遣先住所")
        if data.get("plant_address") or data.get("工場住所"):
            factory.plant_address = data.get("plant_address") or data.get("工場住所")
        if data.get("conflict_date") or data.get("抵触日"):
            cd = data.get("conflict_date") or data.get("抵触日")
            if isinstance(cd, str):
                factory.conflict_date = datetime.strptime(cd, "%Y-%m-%d").date()
            elif isinstance(cd, datetime):
                factory.conflict_date = cd.date()

    # ========================================
    # EMPLOYEE IMPORT
    # ========================================

    def preview_employees_excel(self, content: bytes, sheet_name: str = None) -> ImportResult:
        """
        Preview employee data from Excel file.

        Supports both:
        1. DBGenzaiX format (社員台帳 master file)
        2. Simple template format
        """
        result = ImportResult()

        try:
            # Try to load the workbook to check available sheets
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
            available_sheets = wb.sheetnames
            wb.close()

            # Determine which sheet to use
            target_sheet = sheet_name
            if not target_sheet:
                # Auto-detect: prefer DBGenzaiX if available
                for sheet in ['DBGenzaiX', 'DBGenzai', 'DB現在']:
                    if sheet in available_sheets:
                        target_sheet = sheet
                        break
                if not target_sheet:
                    target_sheet = available_sheets[0]  # Use first sheet

            # Load the sheet
            df = pd.read_excel(BytesIO(content), sheet_name=target_sheet, engine='openpyxl')
            result.total_rows = len(df)

            # Check if this is DBGenzaiX format (has specific columns)
            is_dbgenzai_format = '社員№' in df.columns and '配属ライン' in df.columns

            if is_dbgenzai_format:
                # DBGenzaiX column mapping (社員台帳 master format)
                column_mapping = {
                    '現在': 'status_raw',
                    '社員№': 'employee_number',
                    '派遣先': 'company_name',
                    '配属先': 'department',
                    '配属ライン': 'line_name',
                    '仕事内容': 'position',
                    '氏名': 'full_name_kanji',
                    'カナ': 'full_name_kana',
                    '性別': 'gender',
                    '国籍': 'nationality',
                    '生年月日': 'date_of_birth',
                    '時給': 'hourly_rate',
                    '請求単価': 'billing_rate',
                    'ビザ期限': 'visa_expiry_date',
                    'ビザ種類': 'visa_type',
                    '〒': 'postal_code',
                    '住所': 'address',
                    'アパート': 'apartment_name',
                    'ｱﾊﾟｰﾄ': 'apartment_name',
                    '入社日': 'hire_date',
                    '退社日': 'termination_date',
                    '社保加入': 'insurance_status',
                    '備考': 'notes',
                    '免許種類': 'drivers_license',
                    '免許期限': 'drivers_license_expiry',
                    '日本語検定': 'qualifications',
                }
            else:
                # Simple template format
                column_mapping = {
                    '社員№': 'employee_number',
                    '社員番号': 'employee_number',
                    '氏名': 'full_name_kanji',
                    'カナ': 'full_name_kana',
                    'ローマ字': 'full_name_romaji',
                    '性別': 'gender',
                    '生年月日': 'date_of_birth',
                    '国籍': 'nationality',
                    '住所': 'address',
                    '電話番号': 'phone',
                    '携帯電話': 'mobile',
                    '入社日': 'hire_date',
                    '退社日': 'termination_date',
                    '派遣先': 'company_name',
                    '工場': 'plant_name',
                    '配属先': 'department',
                    'ライン': 'line_name',
                    '時給': 'hourly_rate',
                    '請求単価': 'billing_rate',
                    '在留資格': 'visa_type',
                    'ビザ期限': 'visa_expiry_date',
                    '在留カード番号': 'zairyu_card_number',
                    '雇用保険': 'has_employment_insurance',
                    '健康保険': 'has_health_insurance',
                    '厚生年金': 'has_pension_insurance',
                    '備考': 'notes',
                }

            df = df.rename(columns=column_mapping)

            for idx, row in df.iterrows():
                row_num = idx + 2
                emp_data = row.to_dict()
                emp_data = {k: (None if pd.isna(v) else v) for k, v in emp_data.items()}

                # Convert status_raw to status for DBGenzaiX format
                if is_dbgenzai_format and 'status_raw' in emp_data:
                    status_raw = emp_data.get('status_raw')
                    if status_raw:
                        status_str = str(status_raw).strip().lower()
                        if '退社' in status_str:
                            emp_data['status'] = 'resigned'
                        elif '休職' in status_str:
                            emp_data['status'] = 'on_leave'
                        else:
                            emp_data['status'] = 'active'
                    else:
                        emp_data['status'] = 'active'

                # Convert insurance_status to individual flags
                if is_dbgenzai_format and 'insurance_status' in emp_data:
                    insurance = emp_data.get('insurance_status')
                    if insurance and str(insurance).upper() == 'OK':
                        emp_data['has_health_insurance'] = True
                        emp_data['has_pension_insurance'] = True
                        emp_data['has_employment_insurance'] = True

                validated, errors = self._validate_employee(row_num, emp_data, is_dbgenzai_format)
                result.errors.extend(errors)

                preview_item = {
                    "row": row_num,
                    "employee_number": str(emp_data.get("employee_number", "")) if emp_data.get("employee_number") else "",
                    "full_name_kanji": emp_data.get("full_name_kanji", ""),
                    "full_name_kana": emp_data.get("full_name_kana", ""),
                    "company_name": emp_data.get("company_name", ""),
                    "hourly_rate": emp_data.get("hourly_rate"),
                    "hire_date": str(emp_data.get("hire_date", "")) if emp_data.get("hire_date") else "",
                    "is_valid": len([e for e in errors if e.row == row_num]) == 0,
                    "errors": [e.message for e in errors if e.row == row_num],
                    "_raw": emp_data
                }
                result.preview_data.append(preview_item)

            valid_count = len([p for p in result.preview_data if p["is_valid"]])
            result.success = valid_count > 0  # Success if at least some records are valid
            result.message = f"シート「{target_sheet}」から {result.total_rows}件を読み込みました（有効: {valid_count}件）"

        except Exception as e:
            import traceback
            result.errors.append(ImportValidationError(0, "file", f"Excelエラー: {str(e)}"))
            result.message = f"Excelファイル読み込みエラー: {str(e)}"

        return result

    def import_employees(self, preview_data: List[dict], mode: str = "sync") -> ImportResult:
        """
        Import employees after preview confirmation.

        Args:
            preview_data: Validated data from preview
            mode: "create", "update", or "sync" (update existing, create new)
        """
        result = ImportResult()
        result.total_rows = len(preview_data)

        # Pre-process duplicates: consolidate by employee_number, prioritizing active status
        # This handles Excel files where the same employee appears multiple times
        # (e.g., once as 退社 and once as 在職中)
        consolidated_data: dict = {}
        duplicate_count = 0

        for item in preview_data:
            if not item.get("is_valid", False):
                result.skipped_count += 1
                continue

            raw_data = item.get("_raw", {})
            employee_number = str(raw_data.get("employee_number", "")).strip()

            if not employee_number or employee_number == "0":
                result.skipped_count += 1
                continue

            current_status = raw_data.get("status", "active")

            if employee_number in consolidated_data:
                # Duplicate found - keep the LAST record (highest row number = most recent)
                # This handles employee history: entry -> exit -> re-entry at different factory
                existing_row = consolidated_data[employee_number].get("row", 0)
                current_row = item.get("row", 0)
                if current_row > existing_row:
                    consolidated_data[employee_number] = item
                duplicate_count += 1
            else:
                consolidated_data[employee_number] = item

        try:
            for employee_number, item in consolidated_data.items():
                raw_data = item.get("_raw", {})

                # Check if exists in database
                existing = self.db.query(Employee).filter(
                    Employee.employee_number == employee_number
                ).first()

                if existing:
                    if mode in ["update", "sync"]:
                        self._update_employee(existing, raw_data)
                        result.updated_count += 1
                    else:
                        result.skipped_count += 1
                else:
                    if mode in ["create", "sync"]:
                        employee = self._create_employee(raw_data)
                        self.db.add(employee)
                        result.imported_count += 1
                    else:
                        result.skipped_count += 1

            self.db.commit()
            result.success = True
            dup_msg = f", 重複統合{duplicate_count}件" if duplicate_count > 0 else ""
            result.message = f"同期完了: 新規{result.imported_count}件, 更新{result.updated_count}件, スキップ{result.skipped_count}件{dup_msg}"

        except Exception as e:
            self.db.rollback()
            result.errors.append(ImportValidationError(0, "import", f"インポートエラー: {str(e)}"))
            result.message = f"インポート失敗: {str(e)}"

        return result

    def _validate_employee(self, row: int, data: dict, is_dbgenzai_format: bool = False) -> Tuple[dict, List[ImportValidationError]]:
        """Validate a single employee record."""
        errors = []

        # Employee number is always required
        emp_num = data.get("employee_number")
        if not emp_num or (isinstance(emp_num, str) and not emp_num.strip()):
            errors.append(ImportValidationError(row, "employee_number", "社員番号は必須です"))

        # Name is required
        if not data.get("full_name_kanji"):
            errors.append(ImportValidationError(row, "full_name_kanji", "氏名は必須です"))

        # Kana - required for template format, optional for DBGenzaiX (will default to kanji)
        if not is_dbgenzai_format and not data.get("full_name_kana"):
            errors.append(ImportValidationError(row, "full_name_kana", "カナは必須です"))

        # Hire date - required for template, optional for DBGenzaiX (will default to today)
        if not is_dbgenzai_format and not data.get("hire_date"):
            errors.append(ImportValidationError(row, "hire_date", "入社日は必須です"))

        # Validate dates - but don't fail on format issues for DBGenzaiX
        # (pandas usually handles Excel dates correctly)
        for date_field in ["hire_date", "termination_date", "date_of_birth", "visa_expiry_date"]:
            if data.get(date_field):
                try:
                    val = data[date_field]
                    if isinstance(val, str):
                        # Try to parse if it's a string
                        datetime.strptime(val, "%Y-%m-%d")
                    # datetime and date objects are already valid
                except ValueError:
                    # Only add error for template format, be lenient for DBGenzaiX
                    if not is_dbgenzai_format:
                        errors.append(ImportValidationError(row, date_field, f"{date_field}の日付形式が正しくありません", data[date_field]))

        return data, errors

    def _create_employee(self, data: dict) -> Employee:
        """Create new Employee from import data."""

        def parse_date(val):
            if val is None:
                return None
            if isinstance(val, date):
                return val
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, str):
                val = val.strip()
                if not val or val in ('0', '00:00:00'):
                    return None
                # Try common date formats
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y']:
                    try:
                        return datetime.strptime(val, fmt).date()
                    except ValueError:
                        continue
            if isinstance(val, (int, float)):
                # Excel serial date
                try:
                    from datetime import timedelta
                    excel_epoch = datetime(1899, 12, 30)
                    return (excel_epoch + timedelta(days=int(val))).date()
                except:
                    pass
            return None

        def parse_bool(val):
            if val is None:
                return True
            if isinstance(val, bool):
                return val
            if isinstance(val, str):
                return val.lower() in ['true', '1', 'yes', 'はい', '○', '◯', 'ok']
            return bool(val)

        def parse_decimal(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                if val == 0:
                    return None
                return Decimal(str(val))
            if isinstance(val, str):
                val = val.strip().replace(',', '')
                if not val or val == '0':
                    return None
                try:
                    return Decimal(val)
                except:
                    return None
            return None

        def clean_string(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                if val == 0:
                    return None
                return str(val)
            s = str(val).strip()
            if s in ('0', ''):
                return None
            return s

        def parse_gender(val):
            if val is None:
                return None
            gender_str = str(val).strip()
            if gender_str in ('男', 'M', 'male'):
                return 'male'
            if gender_str in ('女', 'F', 'female'):
                return 'female'
            return 'other'

        # Get status from data or default to active
        status = data.get("status", "active")

        employee = Employee(
            employee_number=str(data.get("employee_number", "")).strip(),
            full_name_kanji=clean_string(data.get("full_name_kanji")) or "Unknown",
            full_name_kana=clean_string(data.get("full_name_kana")) or clean_string(data.get("full_name_kanji")) or "Unknown",
            full_name_romaji=clean_string(data.get("full_name_romaji")),
            gender=parse_gender(data.get("gender")),
            date_of_birth=parse_date(data.get("date_of_birth")),
            nationality=clean_string(data.get("nationality")) or "ベトナム",
            postal_code=clean_string(data.get("postal_code")),
            address=clean_string(data.get("address")),
            apartment_name=clean_string(data.get("apartment_name")),
            phone=clean_string(data.get("phone")),
            mobile=clean_string(data.get("mobile")),
            hire_date=parse_date(data.get("hire_date")) or date.today(),
            termination_date=parse_date(data.get("termination_date")),
            company_name=clean_string(data.get("company_name")),
            plant_name=clean_string(data.get("plant_name")),
            department=clean_string(data.get("department")),
            line_name=clean_string(data.get("line_name")),
            position=clean_string(data.get("position")),
            hourly_rate=parse_decimal(data.get("hourly_rate")),
            billing_rate=parse_decimal(data.get("billing_rate")),
            visa_type=clean_string(data.get("visa_type")),
            visa_expiry_date=parse_date(data.get("visa_expiry_date")),
            zairyu_card_number=clean_string(data.get("zairyu_card_number")),
            has_employment_insurance=parse_bool(data.get("has_employment_insurance")),
            has_health_insurance=parse_bool(data.get("has_health_insurance")),
            has_pension_insurance=parse_bool(data.get("has_pension_insurance")),
            drivers_license=clean_string(data.get("drivers_license")),
            drivers_license_expiry=parse_date(data.get("drivers_license_expiry")),
            qualifications=clean_string(data.get("qualifications")),
            notes=clean_string(data.get("notes")),
            status=status
        )

        # Auto-link employee to factory and factory_line
        company_name = clean_string(data.get("company_name"))
        department = clean_string(data.get("department"))
        line_name = clean_string(data.get("line_name"))

        if company_name:
            factory_id, factory_line_id = self._link_employee_to_factory(
                employee, company_name, department, line_name
            )
            employee.factory_id = factory_id
            employee.factory_line_id = factory_line_id

        return employee

    def _update_employee(self, employee: Employee, data: dict):
        """Update existing employee with new data."""

        def parse_date(val):
            if val is None:
                return None
            if isinstance(val, date):
                return val
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, str):
                return datetime.strptime(val, "%Y-%m-%d").date()
            return None

        def parse_decimal(val):
            if val is None:
                return None
            try:
                return Decimal(str(val))
            except:
                return None

        # Update fields if provided
        if data.get("full_name_kanji"):
            employee.full_name_kanji = data["full_name_kanji"]
        if data.get("full_name_kana"):
            employee.full_name_kana = data["full_name_kana"]
        if data.get("company_name"):
            employee.company_name = data["company_name"]
        if data.get("plant_name"):
            employee.plant_name = data["plant_name"]
        if data.get("department"):
            employee.department = data["department"]
        if data.get("line_name"):
            employee.line_name = data["line_name"]
        if data.get("hourly_rate"):
            employee.hourly_rate = parse_decimal(data["hourly_rate"])
        if data.get("billing_rate"):
            employee.billing_rate = parse_decimal(data["billing_rate"])
        if data.get("visa_expiry_date"):
            employee.visa_expiry_date = parse_date(data["visa_expiry_date"])
        if data.get("termination_date"):
            employee.termination_date = parse_date(data["termination_date"])
            if employee.termination_date:
                employee.status = "resigned"

        # Auto-link employee to factory and factory_line if assignment changed
        company_name = data.get("company_name") or employee.company_name
        department = data.get("department") or employee.department
        line_name = data.get("line_name") or employee.line_name

        if company_name:
            factory_id, factory_line_id = self._link_employee_to_factory(
                employee, company_name, department, line_name
            )
            employee.factory_id = factory_id
            employee.factory_line_id = factory_line_id

    # ========================================
    # FACTORY BULK IMPORT FROM JSON FILES
    # ========================================

    def import_factory_from_json_structure(self, factory_data: dict, mode: str = "sync") -> Tuple[Optional[Factory], List[str]]:
        """
        Import a single factory from the structured JSON format used in E:\config\factories.

        Expected structure:
        {
            "factory_id": "会社名_工場名",
            "client_company": {
                "name": "...",
                "address": "...",
                "phone": "...",
                "responsible_person": { "department": "...", "name": "...", "phone": "..." },
                "complaint_handler": { "department": "...", "name": "...", "phone": "..." }
            },
            "plant": { "name": "...", "address": "...", "phone": "..." },
            "lines": [...],
            "dispatch_company": { ... },
            "schedule": { ... },
            "payment": { ... },
            "agreement": { ... }
        }
        """
        errors = []

        try:
            # Build factory_id
            factory_id = factory_data.get("factory_id", "")
            if not factory_id:
                client = factory_data.get("client_company", {})
                plant = factory_data.get("plant", {})
                factory_id = f"{client.get('name', '')}_{plant.get('name', '')}".replace(" ", "_")

            if not factory_id or factory_id == "_":
                errors.append("factory_id が生成できません")
                return None, errors

            # Check if exists
            existing = self.db.query(Factory).filter(Factory.factory_id == factory_id).first()

            if existing and mode == "create":
                return existing, [f"工場 '{factory_id}' は既に存在します (スキップ)"]

            # Parse data
            client = factory_data.get("client_company", {})
            plant = factory_data.get("plant", {})
            dispatch = factory_data.get("dispatch_company", {})
            schedule = factory_data.get("schedule", {})
            payment = factory_data.get("payment", {})
            agreement = factory_data.get("agreement", {})

            # Parse dates
            def parse_date_str(val):
                if not val:
                    return None
                if isinstance(val, date):
                    return val
                if isinstance(val, datetime):
                    return val.date()
                if isinstance(val, str):
                    val = val.strip()
                    if not val:
                        return None
                    # Try various formats
                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d']:
                        try:
                            return datetime.strptime(val.split()[0] if ' ' in val else val, fmt.split()[0]).date()
                        except ValueError:
                            continue
                return None

            # Parse responsible persons
            client_resp = client.get("responsible_person", {})
            client_complaint = client.get("complaint_handler", {})
            dispatch_resp = dispatch.get("responsible_person", {})
            dispatch_complaint = dispatch.get("complaint_handler", {})

            if existing:
                # Update existing factory
                factory = existing
            else:
                # Create new factory
                factory = Factory(factory_id=factory_id)
                self.db.add(factory)

            # Set/update all fields
            factory.company_name = client.get("name", "")
            factory.company_address = client.get("address")
            factory.company_phone = client.get("phone")

            factory.client_responsible_department = client_resp.get("department")
            factory.client_responsible_name = client_resp.get("name")
            factory.client_responsible_phone = client_resp.get("phone")

            factory.client_complaint_department = client_complaint.get("department")
            factory.client_complaint_name = client_complaint.get("name")
            factory.client_complaint_phone = client_complaint.get("phone")

            factory.plant_name = plant.get("name", "")
            factory.plant_address = plant.get("address")
            factory.plant_phone = plant.get("phone")

            factory.dispatch_responsible_department = dispatch_resp.get("department")
            factory.dispatch_responsible_name = dispatch_resp.get("name")
            factory.dispatch_responsible_phone = dispatch_resp.get("phone")

            factory.dispatch_complaint_department = dispatch_complaint.get("department")
            factory.dispatch_complaint_name = dispatch_complaint.get("name")
            factory.dispatch_complaint_phone = dispatch_complaint.get("phone")

            # Schedule
            factory.work_hours_description = schedule.get("work_hours")
            factory.break_time_description = schedule.get("break_time")
            factory.calendar_description = schedule.get("calendar")
            factory.overtime_description = schedule.get("overtime_labor")
            factory.holiday_work_description = schedule.get("non_work_day_labor")
            factory.conflict_date = parse_date_str(schedule.get("conflict_date"))

            # Parse time unit
            time_unit = schedule.get("time_unit")
            if time_unit:
                try:
                    factory.time_unit_minutes = Decimal(str(time_unit))
                except:
                    pass

            # Payment
            factory.closing_date = payment.get("closing_date")
            factory.payment_date = payment.get("payment_date")
            factory.bank_account = payment.get("bank_account")
            factory.worker_closing_date = payment.get("worker_closing_date")
            factory.worker_payment_date = payment.get("worker_payment_date")
            factory.worker_calendar = payment.get("worker_calendar")

            # Agreement
            factory.agreement_period = parse_date_str(agreement.get("period"))
            factory.agreement_explainer = agreement.get("explainer")

            factory.is_active = True

            # Now import lines
            lines_data = factory_data.get("lines", [])
            existing_line_ids = set()

            for line_data in lines_data:
                assignment = line_data.get("assignment", {})
                job = line_data.get("job", {})
                supervisor = assignment.get("supervisor", {})

                line_id = line_data.get("line_id", "")
                if not line_id:
                    # Generate line_id
                    line_id = f"{factory_id}_{assignment.get('department', '')}_{assignment.get('line', '')}"

                existing_line_ids.add(line_id)

                # Find or create line
                factory_line = None
                if existing:
                    factory_line = self.db.query(FactoryLine).filter(
                        FactoryLine.factory_id == factory.id,
                        FactoryLine.line_id == line_id
                    ).first()

                if not factory_line:
                    factory_line = FactoryLine(line_id=line_id)
                    factory.lines.append(factory_line)

                factory_line.department = assignment.get("department")
                factory_line.line_name = assignment.get("line")

                factory_line.supervisor_department = supervisor.get("department")
                factory_line.supervisor_name = supervisor.get("name")
                factory_line.supervisor_phone = supervisor.get("phone")

                factory_line.job_description = job.get("description")
                factory_line.job_description_detail = job.get("description2")

                # Parse hourly rate
                hourly_rate = job.get("hourly_rate")
                if hourly_rate is not None:
                    try:
                        factory_line.hourly_rate = Decimal(str(hourly_rate))
                    except:
                        pass

                factory_line.is_active = True

            return factory, errors

        except Exception as e:
            errors.append(f"インポートエラー: {str(e)}")
            return None, errors

    def import_factories_from_folder(self, folder_path: str, mode: str = "sync") -> ImportResult:
        """
        Import all factory JSON files from a folder.

        Args:
            folder_path: Path to folder containing factory JSON files
            mode: "create" (skip existing), "update"/"sync" (update existing)
        """
        import os
        import glob

        result = ImportResult()

        try:
            # Find all JSON files in folder
            json_files = glob.glob(os.path.join(folder_path, "*.json"))

            # Exclude non-factory files
            excluded = ["factory_id_mapping.json", "factories_index.json"]
            json_files = [f for f in json_files if os.path.basename(f) not in excluded]

            result.total_rows = len(json_files)

            for idx, json_file in enumerate(json_files, 1):
                filename = os.path.basename(json_file)

                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        factory_data = json.load(f)

                    factory, errors = self.import_factory_from_json_structure(factory_data, mode)

                    if factory and not errors:
                        if hasattr(factory, 'id') and factory.id:
                            result.updated_count += 1
                        else:
                            result.imported_count += 1

                        result.preview_data.append({
                            "row": idx,
                            "file": filename,
                            "factory_id": factory.factory_id,
                            "company_name": factory.company_name,
                            "plant_name": factory.plant_name,
                            "lines_count": len(factory.lines) if hasattr(factory, 'lines') else 0,
                            "is_valid": True,
                            "errors": []
                        })
                    else:
                        if errors and "スキップ" in errors[0]:
                            result.skipped_count += 1
                        else:
                            result.errors.extend([
                                ImportValidationError(idx, "file", e, filename)
                                for e in errors
                            ])

                        result.preview_data.append({
                            "row": idx,
                            "file": filename,
                            "factory_id": factory_data.get("factory_id", ""),
                            "is_valid": False,
                            "errors": errors
                        })

                except json.JSONDecodeError as e:
                    result.errors.append(ImportValidationError(idx, "file", f"JSON解析エラー: {str(e)}", filename))
                except Exception as e:
                    result.errors.append(ImportValidationError(idx, "file", f"読み込みエラー: {str(e)}", filename))

            self.db.commit()
            result.success = True
            result.message = f"フォルダインポート完了: 新規{result.imported_count}件, 更新{result.updated_count}件, スキップ{result.skipped_count}件"

        except Exception as e:
            self.db.rollback()
            result.errors.append(ImportValidationError(0, "folder", f"フォルダエラー: {str(e)}"))
            result.message = f"インポート失敗: {str(e)}"
            result.success = False

        return result
